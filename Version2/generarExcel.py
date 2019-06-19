import xlrd
import openpyxl
from xlwt import Workbook
from pandas import *
from numpy import *
import random

def generarXLSX():
	wb = Workbook()
	sheet1 = wb.add_sheet('Hoja1', cell_overwrite_ok=True)
	totalSucursales= int(input('Ingrese cantidad de sucursales: '))
	totalArticulos= int(input('Ingrese cantidad de articulos: '))
	i=1
	j=1
	totalRegistros=totalArticulos*totalSucursales
	totalSucursales+=1
	totalArticulos+=1
	numfila=1
	for i in range(1,totalSucursales):
		for j in range(1,totalArticulos):
			sheet1.write(numfila,0,i)
			sheet1.write(numfila,1,j)
			numfila=numfila+1

	wb.save('contSucArt.xlsx')

	contSucArt = read_excel('contSucArt.xlsx', sheet_name = 'Hoja1')
	df=pandas.DataFrame(contSucArt)
	df.to_excel('contSucArt.xlsx',header=False, index=False)

	wb = openpyxl.load_workbook(filename="contSucArt.xlsx")
	sheet1 = wb.worksheets[0]

	for i in range(1,totalRegistros+1):
		sheet1.cell(row=i, column=3).value=random.randint(0,100)

	wb.save('contSucArt.xlsx')

generarXLSX()

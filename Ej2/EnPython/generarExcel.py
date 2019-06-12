import xlrd
import openpyxl
from xlwt import Workbook
from pandas import *
from numpy import *
import random

def generarXLSX():
	wb = Workbook()
	sheet1 = wb.add_sheet('Hoja1')

	totalRegistros = int(input('Ingrese cantidad de registros a generar: '))
	totalSucursales= int(input('Ingrese cantidad de sucursales: '))
	totalArticulos= int(input('Ingrese cantidad de articulos: '))
	for i in range(totalRegistros):
		sheet1.write(i,0,random.randint(1,totalSucursales))
		sheet1.write(i,1,random.randint(1,totalArticulos))

	wb.save('contSucArt.xlsx')

	contSucArt 	= read_excel('contSucArt.xlsx', sheet_name = 'Hoja1')
	df = pandas.DataFrame(contSucArt).drop_duplicates()
	totalRegistros=df.count().get_values()[0]
	print("---------------Se han eliminado filas repetidas. Total de filas: ", totalRegistros)
	df.to_excel('contSucArt.xlsx',header=False, index=False)

	wb = openpyxl.load_workbook(filename="contSucArt.xlsx")
	sheet1 = wb.worksheets[0]

	for i in range(1,totalRegistros+1):
		sheet1.cell(row=i, column=3).value=random.randint(0,100)

	wb.save('contSucArt.xlsx')








